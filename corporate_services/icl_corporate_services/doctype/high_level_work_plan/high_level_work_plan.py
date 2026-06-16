# Copyright (c) 2026, IntelliSOFT Consulting and contributors
# For license information, please see license.txt

import os
import tempfile
import openpyxl
import re
import requests

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils.file_manager import get_file
from frappe.utils.data import getdate
from frappe.integrations.doctype.google_drive.google_drive import get_google_drive_object


class HighLevelWorkPlan(Document):
	pass


def _ensure_key_deliverable(value: str):
	"""Ensure a Project Key Deliverable exists for the given value. Returns the name to use in the link field."""
	value = (value or "").strip()
	if not value:
		return None

	if frappe.db.exists("Project Key Deliverable", value):
		return value
	try:
		kd = frappe.get_doc({"doctype": "Project Key Deliverable", "key_name": value})
		kd.insert(ignore_permissions=True)
		return kd.name
	except Exception:
		frappe.log_error(frappe.get_traceback(), "Failed creating Project Key Deliverable during import")
		return value


def _detect_header_row(rows):
	"""Detect header row index by scanning first 20 rows for known header names.
	Returns the index (0-based). If not found, returns 7 (row 8) by default.
	"""
	max_scan = min(20, len(rows))
	candidates = ["line_item", "key_deliverable", "start_date", "end_date", "expected_outcome", "status", "resources", "comments"]
	for i in range(max_scan):
		row = rows[i]
		if not row:
			continue
		normalized = [ (str(cell) if cell is not None else "").strip().lower().replace(" ", "_") for cell in row ]
		# count how many candidate headers present
		matches = sum(1 for c in candidates if c in normalized)
		if matches >= 2:
			return i
	# fallback to row 7 (index 7) if available
	if len(rows) > 7:
		return 7
	# otherwise return first row
	return 0


@frappe.whitelist()
def import_template(docname: str):
	"""Import High Level Work Plan rows from an attached Excel template.

	Assumptions:
	- The header row is located at Excel row 8 (1-based). Data rows follow after that.
	- Expected header names (case-insensitive) map to fields:
	  line item, key deliverable, start date, end date, expected outcome, status, resources, comments
	"""
	doc = frappe.get_doc("High Level Work Plan", docname)
	file_url = getattr(doc, "template_import", None)
	if not file_url:
		frappe.throw(_("No template attached. Please attach an Excel template in 'Template Import'."))

	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_content = get_file(file_doc.file_url)[1]
	extension = (file_doc.file_name or "").split(".")[-1].lower() if file_doc and file_doc.file_name else ""
	if extension not in ("xlsx", "xls", "xlsm"):
		frappe.throw(_("Unsupported file format. Please upload an Excel file (xlsx)."))

	tmp_path = None
	inserted = 0
	inserted_rows = []
	with tempfile.NamedTemporaryFile(delete=False, suffix=f".{extension}") as tmp:
		tmp.write(file_content)
		tmp_path = tmp.name

	try:
		workbook = openpyxl.load_workbook(tmp_path, data_only=True)
		sheet = workbook.active
		rows = list(sheet.iter_rows(values_only=True))

		# Auto-detect header row (scan the first 20 rows); fallback to row 8 if not found
		hdr_idx = _detect_header_row(rows)
		if hdr_idx is None or hdr_idx >= len(rows):
			frappe.throw(_("Could not detect a valid header row in the Excel file."))

		header_row = rows[hdr_idx]
		headers = [ (h or "").strip().lower().replace(" ", "_") for h in header_row ]

		for row in rows[hdr_idx + 1:]:
			# skip empty rows
			if not row or not any(cell not in (None, "") for cell in row):
				continue

			row_map = {}
			for idx, cell in enumerate(row):
				if idx >= len(headers):
					break
				key = headers[idx]
				if not key:
					continue
				value = cell
				# convert dates to ISO if possible
				if key in ("start_date", "end_date"):
					try:
						value = getdate(value).isoformat() if value else None
					except Exception:
						value = str(value) if value else None
				row_map[key] = value

			# Build the child row dict
			child_row = {
				"line_item": row_map.get("line_item") or "",
				"key_deliverable": _ensure_key_deliverable(row_map.get("key_deliverable")),
				"start_date": row_map.get("start_date"),
				"end_date": row_map.get("end_date"),
				"expected_outcome": row_map.get("expected_outcome"),
				"status": row_map.get("status"),
				"resources": row_map.get("resources"),
				"comments": row_map.get("comments"),
			}

			# Skip rows with no meaningful data to avoid blank inserts
			if not any((v is not None and str(v).strip() != "") for v in child_row.values()):
				continue

			doc.append("high_level_workplan_table", child_row)
			inserted += 1
			inserted_rows.append({"line_item": child_row.get("line_item"), "key_deliverable": child_row.get("key_deliverable")})

		# Save and commit to ensure persistence
		doc.save(ignore_permissions=True)
		frappe.db.commit()
	finally:
		if tmp_path:
			try:
				os.unlink(tmp_path)
			except Exception:
				pass

	return {"inserted": inserted, "sample": inserted_rows[:10]}


@frappe.whitelist()
def fetch_workplan_from_drive(docname: str):
	"""Fetch the workplan file from Google Drive (using the google_drive_link_for_the_workplan field)
	and import rows into the high_level_workplan_table.

	Supports both native Google Sheets (exported to XLSX) and regular files.
	"""
	doc = frappe.get_doc("High Level Work Plan", docname)
	link = (getattr(doc, "google_drive_link_for_the_workplan", None) or "").strip()
	if not link:
		frappe.throw(_("No Google Drive link configured on this document."))

	project_name = getattr(doc, "project", None)
	if project_name and not frappe.has_permission("Project", ptype="read", doc=project_name):
		frappe.throw(_("You do not have permission to access the related Project."), frappe.PermissionError)

	# Extract file id from common Drive URL patterns
	m = re.search(r"/d/([A-Za-z0-9_-]+)", link) or re.search(r"id=([A-Za-z0-9_-]+)", link)
	if not m:
		# Try spreadsheets pattern
		m = re.search(r"spreadsheets/d/([A-Za-z0-9_-]+)", link)

	if not m:
		frappe.throw(_("Could not extract a Drive file id from the provided link."))

	file_id = m.group(1)

	# Get access token
	try:
		drive_service, account = get_google_drive_object()
		access_token = account.get_access_token()
	except Exception:
		frappe.log_error(frappe.get_traceback(), "Google Drive access failed")
		frappe.throw(_("Failed to access Google Drive. Please reconnect Google Drive if necessary."))

	# Decide export/download URL
	is_sheet = "spreadsheets" in link or "/spreadsheets/" in link
	if is_sheet:
		# Export Google Sheet as XLSX
		download_url = f"https://www.googleapis.com/drive/v3/files/{file_id}/export?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	else:
		download_url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"

	headers = {"Authorization": f"Bearer {access_token}"}

	tmp_path = None
	inserted = 0
	inserted_rows = []
	resp = requests.get(download_url, headers=headers, timeout=60)
	if resp.status_code >= 400:
		frappe.log_error(f"Drive download failed ({resp.status_code}): {resp.text}", "fetch_workplan_from_drive")
		frappe.throw(_("Failed to download file from Google Drive. Please verify the link and permissions."))

	# Save bytes to temp file and parse similar to import_template
	with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
		tmp.write(resp.content)
		tmp_path = tmp.name

	try:
		workbook = openpyxl.load_workbook(tmp_path, data_only=True)
		sheet = workbook.active
		rows = list(sheet.iter_rows(values_only=True))

		# Auto-detect header row and start importing after it
		hdr_idx = _detect_header_row(rows)
		if hdr_idx is None or hdr_idx >= len(rows):
			frappe.throw(_("Could not detect a valid header row in the Drive file."))

		header_row = rows[hdr_idx]
		headers = [ (h or "").strip().lower().replace(" ", "_") for h in header_row ]

		for row in rows[hdr_idx + 1:]:
			if not row or not any(cell not in (None, "") for cell in row):
				continue

			row_map = {}
			for idx, cell in enumerate(row):
				if idx >= len(headers):
					break
				key = headers[idx]
				if not key:
					continue
				value = cell
				if key in ("start_date", "end_date"):
					try:
						value = getdate(value).isoformat() if value else None
					except Exception:
						value = str(value) if value else None
				row_map[key] = value

			child_row = {
				"line_item": row_map.get("line_item") or "",
				"key_deliverable": _ensure_key_deliverable(row_map.get("key_deliverable")),
				"start_date": row_map.get("start_date"),
				"end_date": row_map.get("end_date"),
				"expected_outcome": row_map.get("expected_outcome"),
				"status": row_map.get("status"),
				"resources": row_map.get("resources"),
				"comments": row_map.get("comments"),
			}

			# Skip empty rows
			if not any((v is not None and str(v).strip() != "") for v in child_row.values()):
				continue

			doc.append("high_level_workplan_table", child_row)
			inserted += 1
			inserted_rows.append({"line_item": child_row.get("line_item"), "key_deliverable": child_row.get("key_deliverable")})

		# Save and commit
		doc.save(ignore_permissions=True)
		frappe.db.commit()
	finally:
		if tmp_path:
			try:
				os.unlink(tmp_path)
			except Exception:
				pass

	return {"inserted": inserted, "sample": inserted_rows[:10]}


@frappe.whitelist()
def get_plans_for_project(project: str):
	"""Return a small payload with High Level and Detailed Work Plan for the given project.

	This is a server-side helper to avoid client-side filter restrictions when querying by
	linked fields like `project`.
	"""
	if not project:
		frappe.throw(_("Project is required"))

	# Respect permissions
	if not frappe.has_permission("Project", ptype="read", doc=project):
		frappe.throw(_("You do not have permission to access this Project"), frappe.PermissionError)

	# Attempt to load project doc for fallback matching
	try:
		project_doc = frappe.get_doc("Project", project)
	except Exception:
		project_doc = None

	# High Level Work Plan: prefer filtering by 'project' field if the doctype has it
	high_filters = {}
	high_meta = frappe.get_meta("High Level Work Plan")
	if high_meta.has_field("project"):
		high_filters["project"] = project
	elif project_doc and getattr(project_doc, "project_name", None):
		high_filters["project_name"] = project_doc.project_name
	else:
		high_filters["project"] = project

	high = frappe.get_all(
		"High Level Work Plan",
		filters=high_filters,
		fields=[
			"name",
			"entry_type",
			"template_import",
			"project_lead",
			"project_start_date",
			"project_end_date",
			"project_duration",
		],
		limit_page_length=1,
	)

	# If a high-level plan was found, fetch its child rows from the child table
	high_rows = []
	if high:
		try:
			high_rows = frappe.get_all(
				"High Level Work Plan Table",
				filters={"parent": high[0]["name"]},
				fields=["line_item", "key_deliverable", "start_date", "end_date", "expected_outcome", "status", "resources", "comments"],
				order_by="idx",
				limit_page_length=1000,
			)
		except Exception:
			frappe.log_error(frappe.get_traceback(), "get_plans_for_project: fetching high level workplan rows failed")

	# Detailed Work Plan: this doctype may not have a 'project' field; try 'project' then 'project_name'
	detailed = None
	det_meta = frappe.get_meta("Detailed Work Plan")
	if det_meta.has_field("project"):
		det_filters = {"project": project}
	elif det_meta.has_field("project_name") and project_doc and getattr(project_doc, "project_name", None):
		det_filters = {"project_name": project_doc.project_name}
	else:
		det_filters = None

	if det_filters:
		try:
			d = frappe.get_all(
				"Detailed Work Plan",
				filters=det_filters,
				fields=["name"],
				limit_page_length=1,
			)
			if d:
				detailed = d[0]
		except Exception:
			frappe.log_error(frappe.get_traceback(), "get_plans_for_project: Detailed Work Plan query failed")

	return {"high": (high[0] if high else None), "high_rows": high_rows, "detailed": (detailed if detailed else None)}
