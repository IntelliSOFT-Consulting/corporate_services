import frappe
from frappe import _
from frappe.utils.file_manager import save_file

from corporate_services.api.notification.project.risk_log import (
	notify_escalation,
	notify_owner_assigned,
)

CATEGORY_MAP = {
	"risks": {
		"doctype": "Risk Assessment List",
		"child_field": "table_okgd",
		"owner_field": "risk_owner",
		"label_field": "risk",
		"closed_statuses": {"Mitigated", "Closed"},
	},
	"assumptions": {
		"doctype": "Project Assumption List",
		"child_field": "assumptions_table",
		"owner_field": "assumption_owner",
		"label_field": "assumption",
		"closed_statuses": {"Invalidated"},
	},
	"issues": {
		"doctype": "Project Issue List",
		"child_field": "issues_table",
		"owner_field": "issue_owner",
		"label_field": "issue",
		"closed_statuses": {"Resolved"},
	},
	"dependencies": {
		"doctype": "Project Dependency List",
		"child_field": "dependencies_table",
		"owner_field": "dependency_owner",
		"label_field": "dependency",
		"closed_statuses": {"Completed"},
	},
}


def _get_category(category):
	config = CATEGORY_MAP.get(category)
	if not config:
		frappe.throw(_("Unknown risk log category: {0}").format(category))
	return config


def _get_assessment_names(project_name):
	return frappe.get_all(
		"Project Risk Assessment",
		filters={"project": project_name},
		fields=["name"],
		order_by="creation asc",
		pluck="name",
	)


@frappe.whitelist()
def get_project_risk_log(project_name):
	if not project_name:
		frappe.throw(_("Project is required."))

	frappe.get_doc("Project", project_name).check_permission("read")

	assessment_names = _get_assessment_names(project_name)

	rows = {key: [] for key in CATEGORY_MAP}
	for assessment_name in assessment_names:
		doc = frappe.get_doc("Project Risk Assessment", assessment_name)
		for key, config in CATEGORY_MAP.items():
			for row in doc.get(config["child_field"]) or []:
				row_dict = row.as_dict()
				row_dict["assessment"] = assessment_name
				rows[key].append(row_dict)

	counts = {}
	for key, config in CATEGORY_MAP.items():
		category_rows = rows[key]
		closed = config["closed_statuses"]
		open_count = sum(1 for r in category_rows if (r.get("status") or "Open") not in closed)
		escalated_count = sum(1 for r in category_rows if r.get("status") == "Escalated")
		counts[key] = {
			"total": len(category_rows),
			"open": open_count,
			"escalated": escalated_count,
		}

	return {
		"assessment": assessment_names[0] if assessment_names else None,
		"risks": rows["risks"],
		"assumptions": rows["assumptions"],
		"issues": rows["issues"],
		"dependencies": rows["dependencies"],
		"counts": counts,
	}


def _get_or_create_assessment(project_name):
	assessment_names = _get_assessment_names(project_name)
	if assessment_names:
		return frappe.get_doc("Project Risk Assessment", assessment_names[0])

	doc = frappe.new_doc("Project Risk Assessment")
	doc.project = project_name
	doc.insert(ignore_permissions=True)
	return doc


@frappe.whitelist()
def save_risk_log_row(project_name, category, row):
	if not project_name:
		frappe.throw(_("Project is required."))

	frappe.get_doc("Project", project_name).check_permission("write")

	config = _get_category(category)
	row = frappe.parse_json(row) if isinstance(row, str) else (row or {})

	doc = _get_or_create_assessment(project_name)
	child_field = config["child_field"]
	child_list = doc.get(child_field)

	row_name = row.get("name")
	child_row = None
	if row_name:
		for existing in child_list:
			if existing.name == row_name:
				child_row = existing
				break
	if not child_row:
		child_row = doc.append(child_field, {})

	owner_field = config["owner_field"]
	owner_before = child_row.get(owner_field)
	status_before = child_row.get("status")

	meta = frappe.get_meta(config["doctype"])
	skip_fields = {"name", "idx", "parent", "parenttype", "parentfield", "doctype"}
	for fieldname, value in row.items():
		if fieldname in skip_fields:
			continue
		if not meta.has_field(fieldname):
			continue
		child_row.set(fieldname, value)

	doc.save(ignore_permissions=True)
	frappe.db.commit()

	owner_after = child_row.get(owner_field)
	status_after = child_row.get("status")
	row_label = child_row.get(config["label_field"])

	if owner_after and owner_after != owner_before:
		notify_owner_assigned(project_name, category, doc.name, row_label, owner_after)

	if status_after == "Escalated" and status_before != "Escalated":
		notify_escalation(project_name, category, doc.name, row_label)

	result = child_row.as_dict()
	result["assessment"] = doc.name
	return result


def _build_workbook(rows_by_category, project_name):
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill

	wb = Workbook()
	wb.remove(wb.active)

	header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)

	sheet_columns = {
		"risks": [
			("risk", "Risk"),
			("areas_affected", "Areas Affected"),
			("severity", "Severity"),
			("likelihood", "Likelihood"),
			("risk_score", "Score"),
			("risk_impact", "Risk Impact"),
			("risk_owner", "Owner"),
			("status", "Status"),
			("recommended_actions", "Recommended Action(s)"),
		],
		"assumptions": [
			("assumption", "Assumption"),
			("areas_affected", "Areas Affected"),
			("status", "Status"),
			("assumption_owner", "Owner"),
			("impact_if_invalid", "Impact if Invalid"),
			("recommended_actions", "Recommended Action(s)"),
		],
		"issues": [
			("issue", "Issue"),
			("areas_affected", "Areas Affected"),
			("severity", "Severity"),
			("issue_owner", "Owner"),
			("status", "Status"),
			("raised_date", "Raised Date"),
			("resolution", "Resolution"),
		],
		"dependencies": [
			("dependency", "Dependency"),
			("depends_on", "Depends On"),
			("dependency_owner", "Owner"),
			("due_date", "Due Date"),
			("status", "Status"),
			("impact_if_delayed", "Impact if Delayed"),
		],
	}

	sheet_titles = {
		"risks": "Risks",
		"assumptions": "Assumptions",
		"issues": "Issues",
		"dependencies": "Dependencies",
	}

	for key in ("risks", "assumptions", "issues", "dependencies"):
		ws = wb.create_sheet(sheet_titles[key])
		columns = sheet_columns[key]
		ws.append([f"{project_name} - {sheet_titles[key]}"])
		ws.append([label for _fieldname, label in columns])
		for cell in ws[2]:
			cell.fill = header_fill
			cell.font = header_font
		for row in rows_by_category.get(key, []):
			ws.append([row.get(fieldname) for fieldname, _label in columns])
		for column_cells in ws.columns:
			length = max((len(str(c.value)) for c in column_cells if c.value), default=10)
			ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

	return wb


@frappe.whitelist()
def export_risk_log_excel(project_name):
	from io import BytesIO

	data = get_project_risk_log(project_name)
	wb = _build_workbook(data, project_name)

	output = BytesIO()
	wb.save(output)
	output.seek(0)

	file_name = f"{project_name}-Risk-Assessment-Log.xlsx"
	file_doc = save_file(file_name, output.read(), "Project", project_name, is_private=0)
	return file_doc.file_url


def _render_pdf_html(data, project_name):
	sheet_titles = {
		"risks": "Risks",
		"assumptions": "Assumptions",
		"issues": "Issues",
		"dependencies": "Dependencies",
	}
	sheet_columns = {
		"risks": [
			("risk", "Risk"),
			("severity", "Severity"),
			("likelihood", "Likelihood"),
			("risk_score", "Score"),
			("risk_impact", "Impact"),
			("risk_owner", "Owner"),
			("status", "Status"),
		],
		"assumptions": [
			("assumption", "Assumption"),
			("status", "Status"),
			("assumption_owner", "Owner"),
			("impact_if_invalid", "Impact if Invalid"),
		],
		"issues": [
			("issue", "Issue"),
			("severity", "Severity"),
			("status", "Status"),
			("issue_owner", "Owner"),
			("raised_date", "Raised Date"),
		],
		"dependencies": [
			("dependency", "Dependency"),
			("depends_on", "Depends On"),
			("status", "Status"),
			("dependency_owner", "Owner"),
			("due_date", "Due Date"),
		],
	}

	html = [f"<h2>{frappe.utils.escape_html(project_name)} - Risk Assessment Log</h2>"]
	for key in ("risks", "assumptions", "issues", "dependencies"):
		columns = sheet_columns[key]
		html.append(f"<h3>{sheet_titles[key]}</h3>")
		html.append('<table border="1" cellspacing="0" cellpadding="4" style="width:100%;border-collapse:collapse;font-size:11px;">')
		html.append("<tr>" + "".join(f"<th>{label}</th>" for _f, label in columns) + "</tr>")
		for row in data.get(key, []):
			html.append(
				"<tr>"
				+ "".join(f"<td>{frappe.utils.escape_html(str(row.get(f) or '-'))}</td>" for f, _label in columns)
				+ "</tr>"
			)
		html.append("</table>")

	return "".join(html)


@frappe.whitelist()
def export_risk_log_pdf(project_name):
	from frappe.utils.pdf import get_pdf

	data = get_project_risk_log(project_name)
	html = _render_pdf_html(data, project_name)
	pdf_content = get_pdf(html)

	file_name = f"{project_name}-Risk-Assessment-Log.pdf"
	file_doc = save_file(file_name, pdf_content, "Project", project_name, is_private=0)
	return file_doc.file_url
